
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font

# =============================================================================
# DIRECTORIES
# =============================================================================

BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input_files" / "helix_chain"
GRAPH_DIR = BASE_DIR / "graphs"
OUTPUT_DIR = BASE_DIR / "output_files"

INPUT_DIR.mkdir(exist_ok=True)
GRAPH_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

# =============================================================================
# SETTINGS
# =============================================================================

CUTOFF = 0.3  # nm

GROUPS = ["pro_mon", "pro_dim", "mem_mon", "mem_dim"]
RESIDUES = list(range(1, 61))

OUTPUT_FILE = OUTPUT_DIR / "helix_closest_residues_tables.xlsx"

# =============================================================================
# CUSTOM COLORS
# =============================================================================

light_pink = "ff98cb"
pink = "ff69b3"
dark_pink = "ec3499"
darkest_pink = "cb1378"

# =============================================================================
# FUNCTIONS
# =============================================================================

def read_xvg_file(file_path):
    residue_distances = {}

    with open(file_path) as file:
        for line in file:
            line = line.strip()

            if not line:
                continue

            if line.startswith("#") or line.startswith("@"):
                continue

            parts = line.split()

            residue = int(parts[0])
            distance = float(parts[1])

            residue_distances[residue] = distance

    return residue_distances


def color_excel_cells(worksheet, table):
    fill_light = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")
    fill_pink = PatternFill(start_color=pink, end_color=pink, fill_type="solid")
    fill_dark = PatternFill(start_color=dark_pink, end_color=dark_pink, fill_type="solid")
    fill_darkest = PatternFill(start_color=darkest_pink, end_color=darkest_pink, fill_type="solid")

    white_font = Font(color="FFFFFF")
    black_font = Font(color="000000")

    for row_idx in range(2, table.shape[0] + 2):
        for col_idx in range(2, table.shape[1] + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = cell.value

            if value is None:
                continue

            if value < 0.15:
                cell.fill = fill_darkest
                cell.font = white_font

            elif value < 0.20:
                cell.fill = fill_dark
                cell.font = white_font

            elif value < 0.25:
                cell.fill = fill_pink
                cell.font = black_font

            elif value < CUTOFF:
                cell.fill = fill_light
                cell.font = black_font


# =============================================================================
# MAIN SCRIPT
# =============================================================================

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

    for group in GROUPS:
        files = []

        for variant in ["wt", "v8a"]:
            for chain in ["chain_a", "chain_b"]:
                files.extend(sorted(INPUT_DIR.glob(f"{group}_{variant}_repl*_*{chain}_closest_residues.xvg")))

        table = pd.DataFrame(index=RESIDUES)

        for file_path in files:
            residue_distances = read_xvg_file(file_path)

            name = file_path.stem

            # Remove group prefix
            name = name.replace(f"{group}_", "")

            # Remove extra text
            name = name.replace("_helix_a_chain_a_closest_residues", "_a")
            name = name.replace("_helix_b_chain_b_closest_residues", "_b")

            table[name] = [
                residue_distances.get(residue, None)
                for residue in RESIDUES
            ]

        table.index.name = "Residue"

        table.to_excel(writer, sheet_name=group)

        worksheet = writer.sheets[group]
        color_excel_cells(worksheet, table)

print("Done! Created Excel file:")
print(OUTPUT_FILE)